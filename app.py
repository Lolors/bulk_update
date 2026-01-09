import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from io import BytesIO
import re
import zipfile


# =========================================================
# 공통 유틸
# =========================================================
def read_csv_flexible(uploaded_file):
    """
    Streamlit 업로드된 CSV 파일을 인코딩/구분자 자동 탐지해서 읽기.
    """
    if uploaded_file is None:
        raise ValueError("CSV 파일이 없습니다.")

    encodings = ["cp949", "utf-8-sig", "utf-8", "ansi", "utf-16"]

    for enc in encodings:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=enc)
            if df.columns.size > 0:
                return df
        except Exception:
            continue

    # 구분자 자동 추측
    try:
        uploaded_file.seek(0)
        df = pd.read_csv(uploaded_file, sep=None, engine="python")
        if df.columns.size > 0:
            return df
    except Exception:
        pass

    raise ValueError("CSV 파일을 읽을 수 없습니다. (인코딩/구분자 문제일 수 있습니다.)")


def adjust_formula_row(formula: str, old_row: int, new_row: int) -> str:
    """
    수식 문자열에서 셀 주소의 '행 번호'만 old_row -> new_row 로 바꾼다.
    예) =($R418+[@외주수량])-$T418  → =($R419+[@외주수량])-$T419

    구조화 참조([@외주수량], 표1[@외주수량]) 같은 건 건드리지 않는다.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    # 셀주소 패턴만 잡기: $ + A~Z 1~3글자 + old_row + 단어 경계
    pattern = re.compile(rf'(\$?[A-Z]{{1,3}}){old_row}\b')

    def repl(m):
        col = m.group(1)
        return f"{col}{new_row}"

    return pattern.sub(repl, formula)


# =========================================================
# 메인 시트 행/열 관련 유틸 (업데이트 탭)
# =========================================================
def find_main_row(ws_main, part_no, lot_no):
    """
    메인 시트에서 품목코드(B), 로트번호(D)가 일치하는 행을 찾는다.
    없으면 0 리턴.
    데이터는 3행부터 시작한다고 가정.
    """
    max_row = ws_main.max_row
    for r in range(3, max_row + 1):
        v_part = str(ws_main.cell(row=r, column=2).value or "").strip()  # B열
        v_lot = str(ws_main.cell(row=r, column=4).value or "").strip()   # D열
        if v_part == str(part_no).strip() and v_lot == str(lot_no).strip():
            return r
    return 0


def get_template_row(ws_main):
    """
    맨 아래에서 위로 올라가면서 로트번호(D열)가 비어있지 않은 행을
    템플릿 행으로 사용.
    """
    for r in range(ws_main.max_row, 2, -1):
        v_lot = str(ws_main.cell(row=r, column=4).value or "").strip()
        if v_lot != "":
            return r
    return None


def get_drum_col_letters():
    """
    VBA 매핑 그대로: X:CE 구간에서 (용량, 위치, 보유통) × 20
      1번: (X,Y,Z)
      2번: (AA,AB,AC)
      ...
      20번: (CC,CD,CE)
    """
    qtyCols = {
        1: "X",  2: "AA", 3: "AD", 4: "AG", 5: "AJ",
        6: "AM", 7: "AP", 8: "AS", 9: "AV", 10: "AY",
        11: "BB", 12: "BE", 13: "BH", 14: "BK", 15: "BN",
        16: "BQ", 17: "BT", 18: "BW", 19: "BZ", 20: "CC",
    }
    locCols = {
        1: "Y",  2: "AB", 3: "AE", 4: "AH", 5: "AK",
        6: "AN", 7: "AQ", 8: "AT", 9: "AW", 10: "AZ",
        11: "BC", 12: "BF", 13: "BI", 14: "BL", 15: "BO",
        16: "BR", 17: "BU", 18: "BX", 19: "CA", 20: "CD",
    }
    stockCols = {
        1: "Z",  2: "AC", 3: "AF", 4: "AI", 5: "AL",
        6: "AO", 7: "AR", 8: "AU", 9: "AX", 10: "BA",
        11: "BD", 12: "BG", 13: "BJ", 14: "BM", 15: "BP",
        16: "BS", 17: "BV", 18: "BY", 19: "CB", 20: "CE",
    }
    return qtyCols, locCols, stockCols


# =========================================================
# 테이블(표1 등) 범위 확장 (업데이트 탭)
# =========================================================
def extend_tables_for_new_row(ws, template_row, new_row):
    """
    template_row 를 포함하는 모든 테이블의 ref 범위를
    new_row 까지 아래로 확장한다.
    (그래야 새 행도 표 안에 들어가고, [@외주수량] 같은 구조화 참조가 정상 작동)
    """
    try:
        tables = ws.tables.values()  # openpyxl 3.x
    except AttributeError:
        tables = ws._tables          # 예전 버전 fallback

    for tbl in tables:
        ref = tbl.ref  # 예: 'F3:U418'
        min_col, min_row, max_col, max_row = range_boundaries(ref)

        if min_row <= template_row <= max_row:
            if new_row > max_row:
                new_ref = (
                    f"{get_column_letter(min_col)}{min_row}:"
                    f"{get_column_letter(max_col)}{new_row}"
                )
                tbl.ref = new_ref


# =========================================================
# 통 업데이트 (VBA ApplyDrumUpdate 핵심, 업데이트 탭)
# =========================================================
def apply_drum_update_to_main(ws_main, row, drum_no, new_qty, new_loc):
    qtyCols, locCols, stockCols = get_drum_col_letters()

    if drum_no < 1 or drum_no > 20:
        st.warning(f"[경고] 통번호 {drum_no} 는 1~20 범위를 벗어남. 스킵.")
        return

    q_col_letter = qtyCols[drum_no]
    l_col_letter = locCols[drum_no]
    s_col_letter = stockCols[drum_no]

    q_col = column_index_from_string(q_col_letter)
    l_col = column_index_from_string(l_col_letter)
    s_col = column_index_from_string(s_col_letter)

    # 새 위치/용량 입력
    ws_main.cell(row=row, column=l_col, value=new_loc)

    loc_upper = str(new_loc or "").strip().upper()
    new_qty_val = float(new_qty or 0)

    if loc_upper in ("소진", "폐기"):
        ws_main.cell(row=row, column=s_col, value=0)        # 보유통
        ws_main.cell(row=row, column=q_col, value=0)        # 용량도 0
    elif loc_upper == "외주":
        ws_main.cell(row=row, column=s_col, value=0)        # 보유통 0
        ws_main.cell(row=row, column=q_col, value=new_qty_val)
    else:
        ws_main.cell(row=row, column=q_col, value=new_qty_val)
        ws_main.cell(row=row, column=s_col, value=0 if new_qty_val == 0 else 1)


# =========================================================
# LOG 시트 한줄 추가 (업데이트 탭)
# =========================================================
def append_log_row(ws_log, log_row):
    """
    LOG 시트에 한 줄 추가.
    - 헤더를 보고 컬럼 위치를 자동 인식 (ID 유무 자동 대응)
    - CSV(log_row)가 ID를 가지고 있으면 그대로 기록
    - CSV에 ID가 없으면 빈칸(또는 None)으로 둠
    """

    # 1) 헤더 읽어서 "헤더명 -> 열번호" 맵 생성
    header_map = {}
    max_col = ws_log.max_column
    for c in range(1, max_col + 1):
        hv = ws_log.cell(row=1, column=c).value
        if hv is None:
            continue
        header_map[str(hv).strip()] = c

    # 2) 새 행 위치
    last = ws_log.max_row
    if last < 1:
        last = 1
    new_r = last + 1

    # 3) 서식 복사(가능한 범위만)
    src_row = last
    for c in range(1, max_col + 1):
        src_cell = ws_log.cell(row=src_row, column=c)
        dst_cell = ws_log.cell(row=new_r, column=c)
        dst_cell._style = src_cell._style

    # 4) 값 채우기 (CSV 컬럼명과 LOG 헤더명을 매칭)
    # CSV쪽 키: 시간, ID, 품번, 품명, 로트번호, 통번호, 변경 전 용량, 변경 후 용량, 변화량, 변경 전 위치, 변경 후 위치
    value_map = {
        "시간": log_row.get("시간"),
        "ID": log_row.get("ID"),
        "품목코드": log_row.get("품번"),   # 엑셀 매크로 헤더가 "품목코드"인 경우도 대응
        "품번": log_row.get("품번"),
        "품명": log_row.get("품명"),
        "로트번호": log_row.get("로트번호"),
        "통번호": log_row.get("통번호"),
        "변경 전 용량": log_row.get("변경 전 용량"),
        "변경 후 용량": log_row.get("변경 후 용량"),
        "변화량": log_row.get("변화량"),
        "변경 전 위치": log_row.get("변경 전 위치"),
        "변경 후 위치": log_row.get("변경 후 위치"),
    }

    for hdr, val in value_map.items():
        if hdr in header_map:
            ws_log.cell(row=new_r, column=header_map[hdr], value=val)

# =========================================================
# bulk_drums_extended 메타 구축 (업데이트 탭)
# =========================================================
def build_meta_from_extended(file_extended):
    """
    bulk_drums_extended.csv 헤더:
    품목코드, 품명, 로트번호, 제품라인, 제조일자, 상태, 통번호, 통용량, 현재위치
    (품목코드, 로트번호) 별로 메타 정보 생성
    """
    df_ext = read_csv_flexible(file_extended)

    if "제조일자" in df_ext.columns:
        df_ext["제조일자"] = pd.to_datetime(df_ext["제조일자"], errors="coerce")

    meta = {}
    grouped = df_ext.groupby(["품목코드", "로트번호"], dropna=False)

    for (part, lot), grp in grouped:
        part_str = str(part).strip()
        lot_str = str(lot).strip()

        # 제품라인
        product_line_series = grp["제품라인"].dropna().astype(str)
        product_line = product_line_series.iloc[0] if not product_line_series.empty else ""

        # 제조일자
        mfg = grp["제조일자"].dropna()
        mfg_date = mfg.iloc[0] if not mfg.empty else None

        # 전체통수 = 통번호 고유 개수
        total_drums = grp["통번호"].nunique()

        # 품명
        name_series = grp["품명"].dropna().astype(str)
        name_val = name_series.iloc[0] if not name_series.empty else ""

        meta[(part_str, lot_str)] = {
            "제품라인": product_line,
            "제조일자": mfg_date,
            "전체통수": total_drums,
            "품명": name_val,
        }

    return meta


# =========================================================
# 신규 로트 행 생성 (업데이트 탭)
# =========================================================
def create_new_main_row(ws_main, part_no, lot_no, prod_name, meta_map, template_row):
    """
    메인 시트에 신규 로트 행 추가.
    - B: 품목코드(=품번)
    - C: 품명
    - D: 로트번호
    - E: 제품라인 (meta)
    - G: 제조일자 (meta)
    - W: 전체통수 (meta)
    - F,H,I,N,O,P,R,S,T,U,V : 템플릿 행 수식을 복사하되, 행 번호만 old->new 교체
    - 전체 행에 대해서 "위 행" 스타일 복사
    - 새 행이 들어가는 만큼 관련 테이블(ref)도 아래로 확장
    """

    key = (str(part_no).strip(), str(lot_no).strip())
    meta = meta_map.get(key, {})

    new_row = ws_main.max_row + 1

    # 0) 관련 테이블 범위를 new_row까지 확장
    if template_row is not None:
        extend_tables_for_new_row(ws_main, template_row, new_row)

    # 1) 위 행(또는 템플릿 행)의 서식을 새 행에 전체 복사
    base_row = template_row if template_row else new_row - 1
    max_col = ws_main.max_column

    for c in range(1, max_col + 1):
        src = ws_main.cell(row=base_row, column=c)
        dst = ws_main.cell(row=new_row, column=c)
        dst._style = src._style

    # 2) B,C,D,E,G,W 값 채우기
    ws_main[f"B{new_row}"] = str(part_no).strip()
    ws_main[f"C{new_row}"] = prod_name if prod_name else meta.get("품명", "")
    ws_main[f"D{new_row}"] = str(lot_no).strip()
    ws_main[f"E{new_row}"] = meta.get("제품라인", "")
    ws_main[f"G{new_row}"] = meta.get("제조일자", None)
    ws_main[f"W{new_row}"] = meta.get("전체통수", None)

    # 3) F,H,I,N,O,P,R,S,T,U,V 수식/값 복사 + 행 번호 치환
    formula_cols = ["F", "H", "I", "N", "O", "P", "R", "S", "T", "U", "V"]

    if template_row is not None:
        for col in formula_cols:
            src = ws_main[f"{col}{template_row}"]
            dst = ws_main[f"{col}{new_row}"]

            val = src.value

            if isinstance(val, str) and val.startswith("="):
                dst.value = adjust_formula_row(val, template_row, new_row)
            else:
                dst.value = val

    return new_row


# =========================================================
# 메인 처리 (업데이트 탭)
# =========================================================
def process_bulk_log_streamlit(excel_file, bulk_log_file, bulk_ext_file):
    """
    업로드된 파일 객체 3개를 받아서
    엑셀(메인+LOG)에 로그 반영 후 BytesIO 로 반환
    """
    # 0) extended 메타 준비
    meta_map = build_meta_from_extended(bulk_ext_file)

    # 1) 엑셀 로드 (매크로 유지)
    excel_bytes = excel_file.read()
    wb = load_workbook(BytesIO(excel_bytes), keep_vba=True)

    if "메인" not in wb.sheetnames or "LOG" not in wb.sheetnames:
        raise ValueError("엑셀 파일에 '메인' 또는 'LOG' 시트가 없습니다.")

    ws_main = wb["메인"]
    ws_log = wb["LOG"]

    template_row = get_template_row(ws_main)

    # 2) LOG 시트에서 기존 마지막 시간 읽기
    log_times = []
    for r in range(2, ws_log.max_row + 1):
        val = ws_log.cell(row=r, column=1).value
        if isinstance(val, datetime):
            log_times.append(val)
    last_time = max(log_times) if log_times else datetime.min

    # 3) bulk_move_log 읽기
    df_log = read_csv_flexible(bulk_log_file)

    # 필수 컬럼 체크
    required_cols = [
        "시간", "품번", "품명", "로트번호", "통번호",
        "변경 전 용량", "변경 후 용량", "변화량",
        "변경 전 위치", "변경 후 위치",
    ]

    # 필수 컬럼 체크
    for col in required_cols:
        if col not in df_log.columns:
            raise ValueError(f"bulk_move_log.csv 에 '{col}' 컬럼이 없습니다.")

    # ID 컬럼은 선택 (없으면 자동 생성)
    if "ID" not in df_log.columns:
        df_log["ID"] = None

    df_log["시간"] = pd.to_datetime(df_log["시간"], errors="coerce")

    # 4) 신규 로그만 필터 + 정렬
    new_logs = df_log[df_log["시간"] > last_time].copy()
    new_logs = new_logs.sort_values("시간")

    if new_logs.empty:
        st.info("반영할 신규 로그가 없습니다. (bulk_move_log.csv 기준)")
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out, 0

    # 5) 신규 로그 한 줄씩 반영
    applied_count = 0
    for _, row in new_logs.iterrows():
        part_no = str(row["품번"]).strip()
        lot_no = str(row["로트번호"]).strip()
        prod_name = str(row.get("품명", "") or "").strip()
        drum_no = int(row["통번호"])
        new_qty = row["변경 후 용량"]
        new_loc = str(row["변경 후 위치"])

        main_row = find_main_row(ws_main, part_no, lot_no)

        if main_row == 0:
            st.write(f"[신규 로트 생성] 품번={part_no}, 로트={lot_no}")
            main_row = create_new_main_row(
                ws_main,
                part_no=part_no,
                lot_no=lot_no,
                prod_name=prod_name,
                meta_map=meta_map,
                template_row=template_row,
            )
            template_row = main_row

        apply_drum_update_to_main(ws_main, main_row, drum_no, new_qty, new_loc)
        append_log_row(ws_log, row.to_dict())
        applied_count += 1

    # 6) 결과를 BytesIO 로 반환
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out, applied_count


# =========================================================
# 추출 탭: 메인 시트 → bulk_drums_extended.csv
# (기존 app.py 로직 기반)
# =========================================================
def extract_bulk_drums_from_main(excel_bytes: bytes, sheet_name: str = "메인") -> pd.DataFrame:
    """
    메인 엑셀 시트에서 bulk_drums_extended 형식으로 추출.

    규칙:
    1) 세부위치(CF~CY)가 있으면 -> 현재위치 = 세부위치
    2) 세부위치가 없으면 -> 메인 위치(층) 보고
       - 외주/폐기/소진: 그대로
       - 그 외: "X층 보관" 으로 강제
    """

    # (A) pandas로 기본 데이터(품목/로트/통용량/층/보유통 등) 읽기
    df = pd.read_excel(BytesIO(excel_bytes), sheet_name=sheet_name, header=1)

    if "1번" not in df.columns:
        raise ValueError(f'"{sheet_name}" 시트에서 "1번" 컬럼을 찾지 못했습니다.')

    start_idx = df.columns.get_loc("1번")

    # (B) openpyxl로 세부위치(CF~CY) 읽기
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    ws = wb[sheet_name]

    detail_cols = {
        1: "CF", 2: "CG", 3: "CH", 4: "CI", 5: "CJ",
        6: "CK", 7: "CL", 8: "CM", 9: "CN", 10: "CO",
        11: "CP", 12: "CQ", 13: "CR", 14: "CS", 15: "CT",
        16: "CU", 17: "CV", 18: "CW", 19: "CX", 20: "CY",
    }

    def norm_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and pd.isna(v):
            return ""
        s = str(v).strip()
        return s

    def to_floor_str(floor_raw: str) -> str:
        """
        '4' -> '4층'
        '4층' -> '4층'
        '4층 보관' -> '4층'
        '창고 보관' -> '창고'
        """
        s = norm_str(floor_raw)
        if not s:
            return ""
        # ✅ 특수 구역은 뒤 텍스트가 붙어도 그대로 유지 (예: '창고 보관' → '창고')
        for special in ("외주", "폐기", "소진", "창고"):
            if s == special or s.startswith(f"{special} ") or s.startswith(f"{special}_") or s.startswith(f"{special}-"):
                return special
        if re.fullmatch(r"\d+", s):
            return f"{s}층"
        m = re.search(r"(\d+층)", s)
        return m.group(1) if m else ""

    rows = []

    # DataFrame의 첫 행(index 0) = 엑셀 3행 (1행 제목, 2행 헤더라는 전제)
    for excel_row, (_, row) in enumerate(df.iterrows(), start=3):

        part = row.get("품목코드")
        name = row.get("품명")

        # 품목 정보 없으면 스킵
        if (pd.isna(part) or norm_str(part) == "") and (pd.isna(name) or norm_str(name) == ""):
            continue

        base = {
            "품목코드": norm_str(part),
            "품명": norm_str(name),
            "로트번호": norm_str(row.get("로트번호")),
            "제품라인": norm_str(row.get("제품라인")),
            "제조일자": pd.to_datetime(row.get("제조일자")).date()
                        if not pd.isna(row.get("제조일자")) else pd.NaT,
            "상태": norm_str(row.get("상태")),
        }

        for drum_no in range(1, 21):
            cap_col = start_idx + 3 * (drum_no - 1)      # 통용량
            loc_col = start_idx + 3 * (drum_no - 1) + 1  # 위치(층)
            cnt_col = start_idx + 3 * (drum_no - 1) + 2  # 보유통

            cap = row.iloc[cap_col]
            floor_raw = row.iloc[loc_col]
            cnt_raw = row.iloc[cnt_col]

            # 1) 세부위치(숨김열) 먼저 읽기
            detail_cell = f"{detail_cols[drum_no]}{excel_row}"
            detail_val = norm_str(ws[detail_cell].value)

            # ✅ 정규화: 세부위치에 'n층 보관' 같은 값이 들어오면
            #    "세부위치가 있는 것"으로 보지 않고, 층 기반 로직으로 보내기
            if detail_val:
                # '보관'이 중복된 경우(보관 보관) 1개로 축약
                detail_val = re.sub(r"(보관)(\s+\1)+", r"\1", detail_val).strip()

                # 세부위치가 사실상 없다는 의미의 값들은 비워서(=층 기준 처리)
            # 2) 현재위치 결정
            floor = to_floor_str(floor_raw)
            detail = norm_str(detail_val)

            # 🔒 세부위치 정리
            if detail in ("", "보관", f"{floor} 보관"):
                detail = ""

            # ✅ 특수 위치는 단독 처리
            if floor in ("외주", "폐기", "소진", "창고"):
                current_loc = floor

            else:
                # 일반 층 (예: 4층, 5층 ...)
                if not floor:
                    current_loc = detail or ""
                else:
                    # 세부위치에 이미 '4층' 같은 게 붙어 있으면 제거
                    if detail.startswith(floor):
                        detail = detail[len(floor):].strip()

                    if not detail:
                        detail = "보관"

                    current_loc = f"{floor} {detail}"


            # 통수량 파싱
            n = 0
            s_cnt = norm_str(cnt_raw)
            if s_cnt:
                try:
                    n = int(float(s_cnt))
                except Exception:
                    n = 0

            # 외주/폐기/소진인데 통수량이 비어있으면 1개로 간주(기존 규칙 유지)
            if current_loc in ("외주", "폐기", "소진") and n <= 0:
                n = 1

            # 용량/위치/통수량 모두 의미 없으면 스킵
            cap_empty = (pd.isna(cap) or cap == 0)
            if n <= 0 and cap_empty and not current_loc:
                continue

            if n <= 0:
                continue

            # 통수량만큼 행 생성
            for _ in range(n):
                rows.append({
                    **base,
                    "통번호": drum_no,
                    "통용량": cap,
                    "현재위치": current_loc,
                })

    return pd.DataFrame(rows)

# =========================================================
# 추출 탭: 시트별로 파일 만들기 + ZIP 묶기
# =========================================================
def extract_and_zip(excel_file):
    """
    벌크 관리대장 엑셀 업로드 파일로부터:
      1) bulk_drums_extended.csv
      2) 제조작업실적현황 → production.xlsx
      3) 일자별통합재고현황 → stock.xlsx
      4) 입하현황 → receive.xlsx
      5) LOG → bulk_move_log.csv (UTF-8)
    를 하나의 ZIP으로 묶어 BytesIO 반환
    """
    excel_bytes = excel_file.read()

    # 1) bulk_drums_extended.csv (메인 시트 기반)
    df_drums = extract_bulk_drums_from_main(excel_bytes)
    drums_buf = BytesIO()
    df_drums.to_csv(drums_buf, index=False, encoding="utf-8-sig")
    drums_bytes = drums_buf.getvalue()

    # 2~4,5는 pandas로 각각 읽어서 내보내기
    # 여러 번 읽어야 하므로 BytesIO 새로 만들어 사용
    excel_buf_for_pd = BytesIO(excel_bytes)

    sheet_names = {
        "제조작업실적현황": "production.xlsx",
        "일자별통합재고현황": "stock.xlsx",
        "입하현황": "receive.xlsx",
    }

    xlsx_files = {}
    for sheet, fname in sheet_names.items():
        excel_buf_for_pd.seek(0)
        try:
            df_sheet = pd.read_excel(excel_buf_for_pd, sheet_name=sheet)
        except ValueError:
            # 시트가 없는 경우 스킵
            continue

        out_buf = BytesIO()
        # engine='openpyxl'은 기본이라 생략 가능
        df_sheet.to_excel(out_buf, index=False)
        xlsx_files[fname] = out_buf.getvalue()

    # 5) LOG 시트 → bulk_move_log.csv (UTF-8, ID 포함 표준화)
    excel_buf_for_pd.seek(0)
    try:
        df_log = pd.read_excel(excel_buf_for_pd, sheet_name="LOG")

        # 컬럼명 정리
        df_log.columns = [str(c).strip() for c in df_log.columns]

        # "시간"은 반드시 있어야 의미 있음. 없으면 그대로 내보내되, 아래 표준컬럼은 만들 수 있는 만큼 만든다.
        if "ID" not in df_log.columns:
            df_log.insert(1, "ID", "")

        # 표준 컬럼 순서 강제 (없는 컬럼은 빈 값으로 생성)
        std_cols = [
            "시간", "ID", "품번", "품명", "로트번호", "통번호",
            "변경 전 용량", "변경 후 용량", "변화량",
            "변경 전 위치", "변경 후 위치",
        ]

        # "품목코드"로 저장된 경우도 "품번"으로 맞춰주기
        if "품번" not in df_log.columns and "품목코드" in df_log.columns:
            df_log["품번"] = df_log["품목코드"]

        for c in std_cols:
            if c not in df_log.columns:
                df_log[c] = ""

        df_log = df_log[std_cols]

        log_buf = BytesIO()
        df_log.to_csv(log_buf, index=False, encoding="utf-8-sig")
        log_bytes = log_buf.getvalue()
    except ValueError:
        log_bytes = b""

    # ZIP 묶기
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("bulk_drums_extended.csv", drums_bytes)

        for fname, data in xlsx_files.items():
            zf.writestr(fname, data)

        zf.writestr("bulk_move_log.csv", log_bytes)

    zip_buf.seek(0)
    return zip_buf


# =========================================================
# Streamlit UI
# =========================================================
def main():
    st.title("벌크 관리대장 도우미")

    tab_update, tab_extract = st.tabs(["업데이트", "추출"])

    # ---------------------- 업데이트 탭 ----------------------
    with tab_update:
        st.subheader("LOG 기준 메인/LOG 업데이트")

        excel_file = st.file_uploader(
            "1) 벌크 관리대장 엑셀 (.xlsm)", type=["xlsm", "xlsx"], key="upd_excel"
        )
        bulk_log_file = st.file_uploader(
            "2) bulk_move_log.csv", type=["csv"], key="upd_log"
        )
        bulk_ext_file = st.file_uploader(
            "3) bulk_drums_extended.csv", type=["csv"], key="upd_ext"
        )

        if st.button("로그 반영 실행", key="run_update"):
            if not excel_file or not bulk_log_file or not bulk_ext_file:
                st.error("세 파일 모두 업로드 해주세요.")
            else:
                try:
                    updated_bytes, applied_count = process_bulk_log_streamlit(
                        excel_file, bulk_log_file, bulk_ext_file
                    )
                except Exception as e:
                    st.error(f"처리 중 오류가 발생했습니다: {e}")
                else:
                    st.success(f"처리 완료! 신규 로그 {applied_count}건이 반영되었습니다.")
                    st.download_button(
                        label="수정된 벌크 관리대장 다운로드",
                        data=updated_bytes,
                        file_name="벌크 관리대장_로그반영.xlsm",
                        mime="application/vnd.ms-excel",
                        key="upd_download",
                    )

    # ---------------------- 추출 탭 ----------------------
    with tab_extract:
        st.subheader("벌크 관리대장에서 세트 파일 추출")

        excel_extract_file = st.file_uploader(
            "벌크 관리대장 엑셀 (.xlsm) 업로드", type=["xlsm", "xlsx"], key="ext_excel"
        )

        if st.button("세트 파일 추출", key="run_extract"):
            if not excel_extract_file:
                st.error("벌크 관리대장 엑셀 파일을 업로드 해주세요.")
            else:
                try:
                    zip_buf = extract_and_zip(excel_extract_file)
                except Exception as e:
                    st.error(f"추출 중 오류가 발생했습니다: {e}")
                else:
                    st.success("추출 완료! 아래 버튼으로 ZIP 파일을 다운로드하세요.")
                    st.download_button(
                        label="추출 결과 ZIP 다운로드",
                        data=zip_buf.getvalue(),
                        file_name="bulk_bundle_export.zip",
                        mime="application/zip",
                        key="ext_download",
                    )


if __name__ == "__main__":
    main()
